"""
分类结果透视：读取「分类结果」Excel，对指定标签列分别做频次透视。
透视列顺序与分类器输出列一致，来自对应类目 config/<类目>/config.json 中的
column_settings.common_output_columns 与 custom_output_columns（例如投影仪见
config/投影仪/config.json 第 14 行起）。

使用：
  python run.py pivot                          # 在 output/ 下按修改时间取最新（含子目录）
  python run.py pivot output/沙发/xxx.xlsx    # 指定 xlsx（可相对工程目录或绝对路径）
  python run.py pivot 沙发                     # 仅在 output/<类目>/ 下取最新
  python run.py pivot sofa                     # 支持 config/_aliases.json 类目别名

导出：单工作表「透视汇总」，各维度按「行标签列标题 + 计数项:维度 + 占合计百分比 + 总计」纵向排列；
块与块之间空 7 行；表头浅蓝底居中；占比列为数值百分比并加蓝色数据条条件格式（与 Excel 透视类似）。
"""

from __future__ import annotations

import json
import os
import re
from collections import Counter
from datetime import datetime
from typing import List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter

from result_saver import ResultSaver

# 透视表块之间的空行数（不含块内数据行）
_PIVOT_BLOCK_GAP_ROWS = 7

# 表头：浅蓝底（接近 Excel 默认透视表样式）
_PIVOT_HEADER_FILL = PatternFill(
    start_color="B4C6E7", end_color="B4C6E7", fill_type="solid"
)
_PIVOT_HEADER_ALIGN = Alignment(
    horizontal="center", vertical="center", wrap_text=True
)
_CELL_ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
_CELL_ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

# 占比列数据条（显示数值 + 蓝色条）
_PCT_DATABAR_RULE = DataBarRule(
    start_type="min",
    end_type="max",
    color=Color(rgb="FF5B9BD5"),
    showValue=True,
)

def load_pivot_dimension_columns_from_config(config_path: str) -> Tuple[str, ...]:
    """
    与分类器一致：column_settings.common_output_columns + custom_output_columns。
    对应各类目 config.json（如 config/投影仪/config.json 第 14 行起的列表）。
    """
    with open(config_path, "r", encoding="utf-8-sig") as f:
        cfg = json.load(f)
    col = cfg.get("column_settings") or {}
    out: List[str] = []
    for key in ("common_output_columns", "custom_output_columns"):
        for x in col.get(key) or []:
            s = str(x).strip()
            if s:
                out.append(s)
    return tuple(out)


def infer_category_folder_from_source_xlsx(
    project_dir: str, source_xlsx: str
) -> Optional[str]:
    """
    若分类结果位于 output/<类目名>/xxx.xlsx，且存在 config/<类目名>/config.json，则返回类目文件夹名。
    """
    proj = os.path.abspath(project_dir)
    src = os.path.abspath(source_xlsx)
    output_root = os.path.join(proj, "output")
    try:
        rel = os.path.relpath(src, output_root)
    except ValueError:
        return None
    if rel.startswith(".."):
        return None
    parts = rel.split(os.sep)
    if len(parts) < 2:
        return None
    folder = parts[0]
    cfg_path = os.path.join(proj, "config", folder, "config.json")
    if folder and os.path.isfile(cfg_path):
        return folder
    return None


def resolve_pivot_dimension_columns(
    project_dir: str, source_xlsx: str, category_id: Optional[str] = None
) -> Tuple[str, ...]:
    """
    解析透视列顺序：优先 CLI 传入的类目，其次从源文件路径推断 output/<类目>/，最后回退「沙发」类目配置。
    """
    proj = os.path.abspath(project_dir)
    candidates: List[str] = []
    if category_id and str(category_id).strip():
        candidates.append(str(category_id).strip())
    inferred = infer_category_folder_from_source_xlsx(proj, source_xlsx)
    if inferred and inferred not in candidates:
        candidates.append(inferred)
    if "沙发" not in candidates:
        sofa_cfg = os.path.join(proj, "config", "沙发", "config.json")
        if os.path.isfile(sofa_cfg):
            candidates.append("沙发")

    for folder in candidates:
        cfg_path = os.path.join(proj, "config", folder, "config.json")
        if not os.path.isfile(cfg_path):
            continue
        dims = load_pivot_dimension_columns_from_config(cfg_path)
        if dims:
            return dims
    return ()

# 多标签单元格分隔（与 classifier MULTI_VALUE_JOIN 一致，并兼容常见竖线）
_MULTI_TAG_SPLIT = re.compile(r"[/｜|]+")

_EXCLUDE_NAME_FRAGMENTS = ("透视",)


def _is_candidate_output(filename: str) -> bool:
    fn = filename.lower()
    if not fn.endswith(".xlsx"):
        return False
    # Excel 打开工作簿时生成的临时锁文件，体积小、常为只读/独占，按 mtime 新易被当成「最新」
    if filename.startswith("~$"):
        return False
    if "分类结果" not in filename:
        return False
    for frag in _EXCLUDE_NAME_FRAGMENTS:
        if frag in filename:
            return False
    return True


def iter_classification_xlsx_paths(
    project_dir: str, category_id: Optional[str] = None
) -> List[str]:
    """列出可能的分类结果路径（去重）。"""
    output_root = os.path.join(project_dir, "output")
    if not os.path.isdir(output_root):
        return []

    dirs: List[str] = []
    if category_id:
        dirs.append(os.path.join(output_root, category_id))
    else:
        dirs.append(output_root)
        for name in sorted(os.listdir(output_root)):
            if name.startswith("."):
                continue
            p = os.path.join(output_root, name)
            if os.path.isdir(p):
                dirs.append(p)

    seen: set[str] = set()
    out: List[str] = []
    for d in dirs:
        if not os.path.isdir(d):
            continue
        try:
            names = os.listdir(d)
        except OSError:
            continue
        for fn in names:
            if not _is_candidate_output(fn):
                continue
            full = os.path.normpath(os.path.join(d, fn))
            if full in seen:
                continue
            seen.add(full)
            out.append(full)
    return out


def find_latest_classification_xlsx(
    project_dir: str, category_id: Optional[str] = None
) -> Optional[str]:
    """按修改时间取最新分类结果 xlsx。"""
    paths = iter_classification_xlsx_paths(project_dir, category_id)
    if not paths:
        return None
    paths.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    return paths[0]


def _expand_cell_tags(value) -> List[str]:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return []
    t = str(value).strip()
    if not t or t.lower() == "nan":
        return []
    parts = [p.strip() for p in _MULTI_TAG_SPLIT.split(t)]
    return [p for p in parts if p]


def pivot_one_column(series: pd.Series) -> Tuple[pd.DataFrame, int, int]:
    """
    单列透视：多标签按分隔符拆开后分别计数。

    Returns:
        (透视表 DataFrame, 非空行数, 标签出现总次数)
    """
    counts: Counter[str] = Counter()
    nonempty_rows = 0
    for v in series:
        tags = _expand_cell_tags(v)
        if not tags:
            continue
        nonempty_rows += 1
        for tag in tags:
            counts[tag] += 1
    total = sum(counts.values())
    rows: List[dict] = []
    for tag, c in counts.most_common():
        pct = round((c / total) * 100, 2) if total else 0.0
        rows.append(
            {
                "标签值": tag,
                "数量": c,
                "占该列标签出现次数比例": pct,
            }
        )
    df = pd.DataFrame(rows)
    if df.empty:
        df = pd.DataFrame(columns=["标签值", "数量", "占该列标签出现次数比例"])
    return df, nonempty_rows, total


def _safe_sheet_name(name: str) -> str:
    # Excel 工作表名 ≤31，且不能包含 : \ / ? * [ ]
    bad = r'[:\\/?*\[\]]'
    s = re.sub(bad, "_", name)
    return s[:31] if len(s) > 31 else s


def _default_pivot_output_path(source_xlsx: str, project_dir: str) -> str:
    """生成透视输出相对路径（与源文件同目录，带序号避免覆盖）。"""
    dir_name = os.path.dirname(os.path.relpath(source_xlsx, project_dir))
    base = os.path.basename(source_xlsx)
    stem, _ = os.path.splitext(base)
    stem = re.sub(r"_分类结果(?:（\d+）)?$", "", stem)
    stem = re.sub(r"（\d+）$", "", stem)
    if dir_name and dir_name != ".":
        return os.path.join(dir_name, f"{stem}_透视汇总（1）.xlsx").replace("\\", "/")
    return f"{stem}_透视汇总（1）.xlsx"


def build_pivot_workbook(
    source_path: str,
    project_dir: str,
    category_id: Optional[str] = None,
) -> str:
    """
    读取分类结果并写出透视文件：单工作表「透视汇总」，各维度按 Excel 透视表样式
    （行标签 / 计数项 / 占合计百分比 + 总计）纵向依次排列；块与块之间空 7 行；
    表头浅蓝底居中；占比数据列（不含「总计」行）为百分比格式并加蓝色数据条。

    Returns:
        写出的 xlsx 绝对路径
    """
    dimension_columns = resolve_pivot_dimension_columns(
        project_dir, source_path, category_id=category_id
    )

    def _match_count(frame: pd.DataFrame) -> int:
        cols = set(str(c).strip() for c in frame.columns)
        return sum(1 for name in dimension_columns if name in cols)

    def _looks_like_grouped_header(frame: pd.DataFrame) -> bool:
        if len(frame.columns) == 0:
            return False
        unnamed = sum(
            1
            for c in frame.columns
            if str(c).strip().lower().startswith("unnamed:")
        )
        # 分组表头首行通常大量 Unnamed 列
        return unnamed >= max(8, int(len(frame.columns) * 0.2))

    df = pd.read_excel(source_path, sheet_name=0, engine="openpyxl")
    if dimension_columns and _looks_like_grouped_header(df):
        matched0 = _match_count(df)
        if matched0 == 0:
            # 兼容“首行是分组标题、次行才是字段名”的导出格式
            df_alt = pd.read_excel(source_path, sheet_name=0, engine="openpyxl", header=1)
            if _match_count(df_alt) >= 1:
                df = df_alt

    row_count = len(df)

    per_column_stats: List[Tuple[str, pd.DataFrame, int, int, bool]] = []
    for name in dimension_columns:
        missing = name not in df.columns
        if missing:
            per_column_stats.append((name, pd.DataFrame(), 0, 0, True))
        else:
            pvt, non_empty, tag_total = pivot_one_column(df[name])
            per_column_stats.append((name, pvt, non_empty, tag_total, False))

    saver = ResultSaver(project_dir)
    rel = _default_pivot_output_path(source_path, project_dir)
    out_abs = saver.get_available_filename(rel)

    wb = Workbook()
    ws = wb.active
    ws.title = _safe_sheet_name("透视汇总")
    header_font = Font(bold=True)
    r = 1

    ws.cell(row=r, column=1, value="源文件")
    ws.cell(row=r, column=2, value=os.path.abspath(source_path))
    r += 1
    ws.cell(row=r, column=1, value="读取时间")
    ws.cell(row=r, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    r += 1
    ws.cell(row=r, column=1, value="数据行数")
    ws.cell(row=r, column=2, value=row_count)
    r += 1
    ws.cell(row=r, column=1, value="透视列数")
    ws.cell(row=r, column=2, value=len(dimension_columns))
    r += 2

    if not dimension_columns:
        ws.cell(row=r, column=1, value="（未配置透视列）")
        ws.cell(
            row=r,
            column=2,
            value="未从 config/<类目>/config.json 解析到 common_output_columns，请检查类目路径或配置。",
        )
        r += 2

    def _style_pivot_header_row(row_idx: int) -> None:
        for col in (1, 2, 3):
            c = ws.cell(row=row_idx, column=col)
            c.font = header_font
            c.fill = _PIVOT_HEADER_FILL
            c.alignment = _PIVOT_HEADER_ALIGN

    for name, pvt, _, tag_total, missing in per_column_stats:
        if missing:
            c1 = ws.cell(row=r, column=1, value=name)
            c1.font = header_font
            c1.fill = _PIVOT_HEADER_FILL
            c1.alignment = _PIVOT_HEADER_ALIGN
            r += 1
            ws.cell(row=r, column=1, value="（源表中不存在该列）")
            r += 1
            r += _PIVOT_BLOCK_GAP_ROWS
            continue

        pct_header = f"计数项:{name}占合计的百分比"
        ws.cell(row=r, column=1, value=name)
        ws.cell(row=r, column=2, value=f"计数项:{name}")
        ws.cell(row=r, column=3, value=pct_header)
        _style_pivot_header_row(r)
        r += 1

        if pvt.empty or tag_total == 0:
            c_a = ws.cell(row=r, column=1, value="(无)")
            c_a.alignment = _CELL_ALIGN_LEFT
            c_b = ws.cell(row=r, column=2, value=0)
            c_b.alignment = _CELL_ALIGN_RIGHT
            c_c = ws.cell(row=r, column=3, value="—")
            c_c.alignment = _CELL_ALIGN_RIGHT
            r += 1
            t1 = ws.cell(row=r, column=1, value="总计")
            t2 = ws.cell(row=r, column=2, value=0)
            t3 = ws.cell(row=r, column=3, value="—")
            for t in (t1, t2, t3):
                t.font = header_font
            t1.alignment = _CELL_ALIGN_LEFT
            t2.alignment = _CELL_ALIGN_RIGHT
            t3.alignment = _CELL_ALIGN_RIGHT
            r += 1
        else:
            data_first = r
            for _, prow in pvt.iterrows():
                ws.cell(row=r, column=1, value=prow["标签值"]).alignment = (
                    _CELL_ALIGN_LEFT
                )
                ws.cell(row=r, column=2, value=int(prow["数量"])).alignment = (
                    _CELL_ALIGN_RIGHT
                )
                pct_val = float(prow["占该列标签出现次数比例"])
                c_pct = ws.cell(row=r, column=3, value=pct_val / 100.0)
                c_pct.number_format = "0.00%"
                c_pct.alignment = _CELL_ALIGN_RIGHT
                r += 1
            data_last = r - 1

            t1 = ws.cell(row=r, column=1, value="总计")
            t2 = ws.cell(row=r, column=2, value=int(tag_total))
            t3 = ws.cell(row=r, column=3, value=1.0)
            for t in (t1, t2, t3):
                t.font = header_font
            t1.alignment = _CELL_ALIGN_LEFT
            t2.alignment = _CELL_ALIGN_RIGHT
            t3.alignment = _CELL_ALIGN_RIGHT
            t3.number_format = "0.00%"
            r += 1

            col_c = get_column_letter(3)
            if data_last >= data_first:
                ws.conditional_formatting.add(
                    f"{col_c}{data_first}:{col_c}{data_last}",
                    _PCT_DATABAR_RULE,
                )

        r += _PIVOT_BLOCK_GAP_ROWS

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 26
    ws.column_dimensions["C"].width = 36

    wb.save(out_abs)
    return out_abs


def try_resolve_pivot_input_as_xlsx(project_dir: str, token: str) -> Optional[str]:
    """
    若 token 能解析为已存在的 .xlsx 文件，返回绝对路径；否则返回 None。
    相对路径会依次尝试：相对工程目录、相对当前工作目录。
    """
    raw = str(token).strip().strip('"').strip("'")
    if not raw:
        return None
    if not raw.lower().endswith(".xlsx"):
        return None

    candidates: List[str] = []
    if os.path.isabs(raw):
        candidates.append(os.path.normpath(raw))
    else:
        candidates.append(os.path.normpath(os.path.join(project_dir, raw)))
        candidates.append(os.path.normpath(os.path.join(os.getcwd(), raw)))
        candidates.append(os.path.normpath(raw))

    seen: set[str] = set()
    for c in candidates:
        if c in seen:
            continue
        seen.add(c)
        try:
            if os.path.isfile(c):
                return os.path.abspath(c)
        except OSError:
            continue
    return None


def run_pivot_cli(
    project_dir: str,
    source_xlsx: Optional[str] = None,
    category_id: Optional[str] = None,
) -> str:
    """
    CLI 入口：用指定 Excel，或在 output 下按类目/全局取最新分类结果，生成透视表。

    Args:
        project_dir: 工程根目录
        source_xlsx: 已解析好的分类结果 xlsx 绝对路径；与 category_id 互斥（优先用此路径）
        category_id: 若为 None 且在未指定 source_xlsx 时，在全部 output 子目录中找最新；
            若指定则仅在 output/<category_id>/ 下找最新

    Returns:
        输出文件绝对路径
    """
    if source_xlsx:
        src = os.path.abspath(source_xlsx)
        if not os.path.isfile(src):
            raise FileNotFoundError(f"指定的文件不存在：{src}")
        if not src.lower().endswith(".xlsx"):
            raise ValueError("透视输入须为 .xlsx 文件")
    else:
        src = find_latest_classification_xlsx(project_dir, category_id)
        if not src:
            hint = (
                f"目录「output/{category_id}」"
                if category_id
                else "「output」及其子文件夹"
            )
            raise FileNotFoundError(
                f"未在 {hint} 下找到文件名包含「分类结果」的 xlsx（已排除含「透视」的输出）。"
                f"请先运行分类，或使用：python run.py pivot <文件路径.xlsx>"
            )
    out = build_pivot_workbook(src, project_dir, category_id=category_id)
    print(f"✅ 透视完成\n  源：{src}\n  输出：{out}")
    return out
