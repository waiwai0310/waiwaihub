"""
结果保存模块
负责保存分类结果到 Excel 文件

开发约定：自动化/手工自检请写入 tempfile 或自定义目录，勿向 output/ 写入测试用 xlsx
（避免与正式分类结果混淆）。参见 tests/test_result_saver.py。
"""

import os
import re
import pandas as pd
from typing import Optional, List, Tuple

try:
    from openpyxl.utils import get_column_letter
except ImportError:
    get_column_letter = None


class ResultSaver:
    """结果保存器"""
    
    def __init__(self, base_path: str = None):
        """
        初始化结果保存器
        
        Args:
            base_path: 输出文件的基础路径，如果为 None 则使用本文件所在目录
        """
        if base_path is None:
            base_path = os.path.dirname(os.path.abspath(__file__))
        
        self.base_path = base_path
    
    def get_available_filename(self, base_filename: str) -> str:
        """
        自动生成不重复的文件名
        如果文件已存在，则添加序号
        
        Args:
            base_filename: 基础文件名 (如 "results.xlsx")
        
        Returns:
            可用的文件名
        
        Example:
            如果 results.xlsx 存在，返回 results(1).xlsx
            如果 results(1).xlsx 也存在，返回 results(2).xlsx
        """
        base_path = os.path.join(self.base_path, base_filename)
        base, ext = os.path.splitext(base_path)

        # 兼容已带序号的文件名：xxx（1）.xlsx -> xxx（2）.xlsx
        match = re.search(r"^(.*)（(\d+)）$", base)
        if match:
            stem = match.group(1)
            i = int(match.group(2))
            candidate = f"{stem}（{i}）{ext}"
            while os.path.exists(candidate):
                i += 1
                candidate = f"{stem}（{i}）{ext}"
            return candidate

        # 默认从（1）开始
        i = 1
        candidate = f"{base}（{i}）{ext}"
        while os.path.exists(candidate):
            i += 1
            candidate = f"{base}（{i}）{ext}"
        return candidate
    
    def save_result(
        self,
        df: pd.DataFrame,
        output_filename: str,
        remove_temp_cols: bool = True,
        freeze_header_and_filter: bool = True,
        column_groups: Optional[List[Tuple[str, List[str]]]] = None,
    ) -> str:
        """
        保存分类结果到 Excel 文件
        
        Args:
            df: 包含分类结果的 DataFrame
            output_filename: 输出文件名
            remove_temp_cols: 是否删除临时列（如 __cleaned_text__）
            freeze_header_and_filter: 是否冻结首行并开启筛选（需 openpyxl）
        
        Returns:
            保存的文件完整路径
        
        Raises:
            Exception: 保存失败
        """
        try:
            df = df.copy()
            
            # 删除临时列
            if remove_temp_cols:
                temp_cols = [col for col in df.columns if col.startswith('__')]
                df.drop(columns=temp_cols, errors='ignore', inplace=True)
            
            # 获取不重复的文件名
            output_path = self.get_available_filename(output_filename)
            
            # 如果输出路径包含子目录（如 output/xxx.xlsx），先确保目录存在
            output_dir = os.path.dirname(output_path)
            if output_dir:
                os.makedirs(output_dir, exist_ok=True)
            
            # 保存文件（openpyxl 便于后续设置冻结窗格与筛选）
            df.to_excel(output_path, index=False, engine="openpyxl")

            header_row = 1
            if column_groups:
                try:
                    self._insert_group_header_row(output_path, column_groups)
                    header_row = 2
                except Exception:
                    # 分组头失败不影响主文件
                    pass

            if freeze_header_and_filter:
                try:
                    self._apply_freeze_header_and_autofilter(output_path, header_row=header_row)
                except Exception:
                    # 视图增强失败不影响主文件已写出
                    pass
            
            try:
                print(f"✅ 结果已保存: {output_path}")
            except UnicodeEncodeError:
                print(f"[OK] 结果已保存: {output_path}")
            
            return output_path
        
        except Exception as e:
            raise Exception(f"保存文件失败: {e}")
    
    def get_output_size(self, file_path: str) -> str:
        """
        获取文件大小
        
        Args:
            file_path: 文件路径
        
        Returns:
            格式化的文件大小 (如 "1.5 MB")
        """
        if not os.path.exists(file_path):
            return "未知"
        
        size = os.path.getsize(file_path)
        
        # 转换为更易读的格式
        for unit in ['B', 'KB', 'MB', 'GB']:
            if size < 1024:
                return f"{size:.1f} {unit}"
            size /= 1024
        
        return f"{size:.1f} TB"

    @staticmethod
    def _apply_freeze_header_and_autofilter(xlsx_path: str, header_row: int = 1) -> None:
        """
        冻结首行表头，并对整张数据区域开启 Excel 自动筛选。
        """
        if get_column_letter is None:
            return
        try:
            from openpyxl import load_workbook
        except ImportError:
            return

        wb = load_workbook(xlsx_path)
        try:
            ws = wb.active
            ws.freeze_panes = f"A{header_row + 1}"

            max_row = ws.max_row or 1
            max_col = ws.max_column or 1
            if max_row >= 1 and max_col >= 1:
                last = f"{get_column_letter(max_col)}{max_row}"
                ws.auto_filter.ref = f"A{header_row}:{last}"

            wb.save(xlsx_path)
        finally:
            wb.close()

    @staticmethod
    def _insert_group_header_row(
        xlsx_path: str,
        column_groups: List[Tuple[str, List[str]]],
    ) -> None:
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Alignment, Font, PatternFill
        except ImportError:
            return
        if get_column_letter is None:
            return

        wb = load_workbook(xlsx_path)
        try:
            ws = wb.active

            # 原始表头在第1行；插入后下移到第2行
            ws.insert_rows(1)

            # 读取第2行表头 -> 列号
            header_to_col = {}
            for col in range(1, (ws.max_column or 1) + 1):
                val = ws.cell(row=2, column=col).value
                if isinstance(val, str) and val.strip():
                    header_to_col[val.strip()] = col

            # 写入分组标签并合并同组区间
            palette = [
                "DDEBF7",  # blue
                "E2F0D9",  # green
                "FCE4D6",  # orange
                "FFF2CC",  # yellow
                "E4DFEC",  # purple
                "D9E1F2",  # indigo
                "F8CBAD",  # peach
                "E2EFDA",  # mint
                "FCE4EC",  # pink
            ]
            for i, (group_label, cols) in enumerate(column_groups):
                idxs = [header_to_col[c] for c in cols if c in header_to_col]
                if not idxs:
                    continue
                start = min(idxs)
                end = max(idxs)
                fill = PatternFill(fill_type="solid", fgColor=palette[i % len(palette)])
                top_cell = ws.cell(row=1, column=start, value=group_label)
                top_cell.font = Font(bold=True)
                top_cell.alignment = Alignment(horizontal="center", vertical="center")
                top_cell.fill = fill
                if end > start:
                    ws.merge_cells(start_row=1, start_column=start, end_row=1, end_column=end)
                # 合并区域内所有单元格统一底色
                for c in range(start, end + 1):
                    ws.cell(row=1, column=c).fill = fill

            wb.save(xlsx_path)
        finally:
            wb.close()
