"""
Excel Exporter - Excel表格导出器
用于将爬取的产品数据导出到Excel
"""
import logging
from pathlib import Path
from typing import List, Dict, Any
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class ExcelExporter:
    """Excel导出器"""
    
    # 列定义：(列名, 字段key, 宽度)
    COLUMNS = [
        ('ASIN', 'asin', 12),
        ('首图', 'main_image', 40),
        ('品牌', 'brand', 15),
        ('标题', 'title', 50),
        ('链接', 'link', 40),
        ('价格', 'price', 12),
        ('活动优惠价', 'promo_price', 12),
        ('评分', 'rating', 10),
        ('评论数', 'review_count', 10),
        ('BSR排名', 'bsr_rank', 12),
        ('大类', 'main_category', 25),
        ('大类排名', 'main_category_rank', 12),
        ('小类', 'sub_category', 25),
        ('小类排名', 'sub_category_rank', 12),
        ('上架时间', 'launch_date', 15),
        ('变体数量', 'variant_count', 10),
        ('畅销颜色', 'best_selling_color', 15),
        ('近30天月销量', 'monthly_sales_30d', 12),
        ('最好的月销', 'best_monthly_sales', 12),
        ('卖点1', 'bullet_point_1', 50),
        ('卖点2', 'bullet_point_2', 50),
        ('卖点3', 'bullet_point_3', 50),
        ('卖点4', 'bullet_point_4', 50),
        ('卖点5', 'bullet_point_5', 50),
        ('爬取时间', 'crawl_time', 18),
        ('状态', 'status', 10),
        ('错误信息', 'error', 30),
    ]
    
    def __init__(self, output_dir: Path = None):
        """
        初始化导出器
        
        Args:
            output_dir: 输出目录，默认为data/
        """
        self.output_dir = output_dir or Path('data')
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def _prepare_row_data(self, product: Dict[str, Any]) -> List[Any]:
        """
        准备行数据
        
        Args:
            product: 产品数据字典
        
        Returns:
            行数据列表
        """
        # 处理卖点（bullet points）
        bullet_points = product.get('bullet_points', [])
        for i in range(5):
            key = f'bullet_point_{i+1}'
            product[key] = bullet_points[i] if i < len(bullet_points) else None
        
        # 按列定义顺序提取数据
        row = []
        for _, field_key, _ in self.COLUMNS:
            value = product.get(field_key)
            
            # 处理特殊类型
            if value is None:
                row.append('')
            elif isinstance(value, (list, dict)):
                row.append(str(value))
            else:
                row.append(value)
        
        return row
    
    def _style_header(self, ws):
        """设置表头样式"""
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
    
    def _style_data_cells(self, ws):
        """设置数据单元格样式"""
        # 边框样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 对齐方式
        wrap_alignment = Alignment(vertical='top', wrap_text=True)
        
        # 对所有数据单元格应用样式
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border
                cell.alignment = wrap_alignment
    
    def _set_column_widths(self, ws):
        """设置列宽"""
        for idx, (_, _, width) in enumerate(self.COLUMNS, start=1):
            column_letter = get_column_letter(idx)
            ws.column_dimensions[column_letter].width = width
    
    def _add_data_validation(self, ws):
        """添加数据验证（如有需要）"""
        # 可以在这里添加下拉列表等数据验证
        pass
    
    def export(
        self, 
        products: List[Dict[str, Any]], 
        filename: str = None
    ) -> Path:
        """
        导出产品数据到Excel
        
        Args:
            products: 产品数据列表
            filename: 输出文件名，默认使用时间戳
        
        Returns:
            输出文件路径
        """
        if not products:
            logger.warning("没有数据可导出")
            return None
        
        # 生成文件名
        if not filename:
            timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
            filename = f'amazon_products_{timestamp}.xlsx'
        
        if not filename.endswith('.xlsx'):
            filename += '.xlsx'
        
        output_path = self.output_dir / filename
        
        logger.info(f"开始导出 {len(products)} 个产品到Excel: {output_path}")
        
        try:
            # 创建工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Products'
            
            # 写入表头
            headers = [col_name for col_name, _, _ in self.COLUMNS]
            ws.append(headers)
            
            # 写入数据
            for product in products:
                row_data = self._prepare_row_data(product)
                ws.append(row_data)
            
            # 应用样式
            self._style_header(ws)
            self._set_column_widths(ws)
            self._style_data_cells(ws)
            self._add_data_validation(ws)
            
            # 冻结首行
            ws.freeze_panes = 'A2'
            
            # 保存文件
            wb.save(output_path)
            
            logger.info(f"成功导出到: {output_path}")
            return output_path
            
        except Exception as e:
            logger.exception(f"导出Excel失败: {e}")
            raise
    
    def export_with_summary(
        self,
        products: List[Dict[str, Any]],
        filename: str = None
    ) -> Path:
        """
        导出产品数据并添加汇总统计sheet
        
        Args:
            products: 产品数据列表
            filename: 输出文件名
        
        Returns:
            输出文件路径
        """
        output_path = self.export(products, filename)
        
        if not output_path:
            return None
        
        try:
            # 重新打开文件添加汇总sheet
            wb = openpyxl.load_workbook(output_path)
            
            # 创建汇总sheet
            summary_ws = wb.create_sheet('Summary', 0)
            
            # 统计数据
            total_count = len(products)
            success_count = sum(1 for p in products if p.get('status') == 'success')
            failed_count = total_count - success_count
            
            # 统计有数据的字段
            fields_stats = {}
            for col_name, field_key, _ in self.COLUMNS:
                if field_key not in ['asin', 'link', 'crawl_time', 'status', 'error']:
                    count = sum(1 for p in products if p.get(field_key))
                    fields_stats[col_name] = count
            
            # 写入汇总信息
            summary_ws.append(['Amazon产品数据爬取汇总'])
            summary_ws.append([])
            summary_ws.append(['爬取时间:', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
            summary_ws.append(['总产品数:', total_count])
            summary_ws.append(['成功:', success_count])
            summary_ws.append(['失败:', failed_count])
            summary_ws.append([])
            summary_ws.append(['字段完整度统计:'])
            summary_ws.append(['字段名', '有数据产品数', '完整度'])
            
            for field_name, count in fields_stats.items():
                percentage = f"{count/total_count*100:.1f}%" if total_count > 0 else "0%"
                summary_ws.append([field_name, count, percentage])
            
            # 样式
            title_font = Font(bold=True, size=14)
            summary_ws['A1'].font = title_font
            
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF')
            for cell in summary_ws[9]:
                cell.fill = header_fill
                cell.font = header_font
            
            # 设置列宽
            summary_ws.column_dimensions['A'].width = 25
            summary_ws.column_dimensions['B'].width = 15
            summary_ws.column_dimensions['C'].width = 15
            
            # 保存
            wb.save(output_path)
            logger.info(f"已添加汇总sheet到: {output_path}")
            
        except Exception as e:
            logger.exception(f"添加汇总sheet失败: {e}")
        
        return output_path
